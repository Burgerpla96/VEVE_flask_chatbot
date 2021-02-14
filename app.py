from flask import Flask,render_template,request,jsonify,session
import os
import dialogflow
import uuid #session아이디로 사용

from flask_cors import CORS
#비공개 키를 config파일에 숨겨놓았다.
from settings.config import DIALOG_CONFIG#프로젝트 아이디/api키가 설정된 모듈 import

from selenium import webdriver
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys
import time
from urllib.parse import quote_plus


app = Flask(__name__)
CORS(app)
app.config['JSON_AS_ASCII'] = False
#session을 위한 시크릿 키 - 임의의 문자열
app.secret_key ='fwer!@#njkh%$*hiu'

#환경변수 설정해 애플리케이션 코드에 사용자 인증정보를 제공
#api키를 환경변수에 등록
os.environ['GOOGLE_APPLICATION_CREDENTIALS'] = DIALOG_CONFIG['GOOGLE_APPLICATION_CREDENTIALS']


@app.route('/')
def index():
    session['session_id'] = str(uuid.uuid4())# uuid.uuid4()하면 랜덤한 숫자와 영문의 조합이 나온다.
    #예시) UUID('db2c016c-6309-4ec5-bf82-5316cdb3d7c8')  UUID('e5283a1e-5e57-4d32-b7c5-8915ce9b23a7')
    return render_template('index.html',session=session['session_id'])

@app.route('/message', methods=['POST'])
def handleMessage():#사용자 UI(Client App) 에서 보낸 대화를 받는 메서드
    session['session_id'] = str(uuid.uuid4())#다른 어플리케이션의 ui사용시
    message = request.form['message']
    print('사용자 UI(Client app)에서 입력한 메세지: ',message)
    #project 아이디 얻기
    project_id = DIALOG_CONFIG.get('PROJECT_ID')
    #flask 가 dialogflpow로 부터 받은 응답
    fulfillmentText = response_from_dialogflow(project_id,session['session_id'],message,'ko')

    #dialog로 부터 받은 응답을 클라이언트에 전송
    return jsonify({'message':fulfillmentText})


#dialog로 메세지를 보내기 위한 함수
#즉 위의 webhook() 함수가 실행된다.
#Dialogflow는 텍스트를 처리한 다음 fulfillment 응답을 다시 보낸다.
def response_from_dialogflow(project_id,session_id,message,language_code):
    #step1. dialogflow와 사용자가 상호작용할 세션 클라이언트 생성
    session_client = dialogflow.SessionsClient()
    session_path = session_client.session_path(project_id,session_id)
    #project/프로젝트 아이디/agent/sessions/session아이디로 생성된다.
    print('[session_path]',session_path,sep='\n')
    if message: #사용자가 대화를 입력한 경우, 대화는 utf-8로 인코딩된 처리할 자연어 256자를 넘어서는 안된다.
        #step2. 사용가 메세지(일반 텍스트)로 TexstInput 생성
        text_input = dialogflow.types.TextInput(text=message,language_code=language_code)
        print('[text_input]',text_input,sep='\n')
        '''#TextInput 반환 형식
        text='사용자가 입력한 대화'
        language_code='ko'
        '''
        #step3. 생성된 TextInput 객체로 QueryInput 객체 생성(즉 dialogflow로 전송할 질의 생성)
        query_input = dialogflow.types.QueryInput(text=text_input)
        print('[query_input]', query_input, sep='\n')
        '''# QueryInput 반환 형식
        text {
              text: "사용자 입력 대화"
              language_code: "ko"
            }
        '''
        #step4. DialogFlow로 SessionsClient객체.detect_intent() method로
        #       QueryInput 객체를 보내고 다시 봇 응답(Responses 섹션에 등록한 대화)을 받는다.
        #       즉 A DetectIntentResponse instance 반환
        response = session_client.detect_intent(session=session_path,query_input=query_input)
        print('[response]', response, sep='\n')
        print('[type(response)]', type(response), sep='\n')
        '''
        session_client.detect_intent() 에서
        google.api_core.exceptions.PermissionDenied: 403 IAM permission 에러시

        GOOGLE CONSOLE 로 들어가서 api 사용자 인증정보 -> 하단의 서비스계정 dialogflow선택 -> 우측의 IAM 선택
        -> 추가 버튼으로 .json 파일로 받은 키에서 client_email 값을 소유주로 등록하기
        '''


    return response.query_result.fulfillment_text #dialogflow bot이 응답


#아래 웹 후크용 메서드는 추가적으로 나만의 응답(database에서 읽어오기 등)을 구성하고자 할때 사용,
# 웹 후크를 사용하지 않아도 우리가 dialogflow 에 등록한 사용자 질의문과 응답으로도 충분히 챗봇을 만들 수 있다.

#웹 후크 서비스 : 득 dialogflow 가 인텐트 매칭후
#아래 api 서비스(웹 후크) 를 post로 요청한다.
#전제 조건
#1. 웹 후크를 적용할 인텐트 선택후 fulfillment menu에서 enable 설정
#2. 해당 봇의 좌측 메뉴인 fulfillment tab 에서 아래 url을 등록(localhost 및 http는 불가)

@app.route('/webhook_rpa',methods=['POST'])
def webhook():#fulfillment를 enable로 설덩한
    #dialogflow에서는 json으로 응답을 보낸다.
    webhook_response = request.get_json(force=True)
    print('[webhook_response]', webhook_response, sep='\n')

    program = webhook_response['queryResult']['parameters']['program']
    #noprogram = webhook_response['queryResult']['parameters']['noprogram']

    if '엑셀' in program:
        from openpyxl.workbook import Workbook
        wb = Workbook()
        sheet1 = wb['Sheet']
        sheet1.title = '오늘 할 일'
        sheet1['A1'] = '오늘의 할 일 리스트'
        sheet1.append(['1.','이메일 보내기'])
        sheet1.append(['2.', '고객명단 정리하기'])

        sheet2 = wb.create_sheet('내일 할 일')
        sheet2.cell(row=1,column=1,value='내일의 할 일 리스트')

        #작성내용 엑셀 파일로 저장
        wb.save('todos.xlsx')
        #엑셀 실행
        os.startfile('todos.xlsx')
        reply = {'fulfillmentText':'엑셀 실행합니다.'}#dialogFlow에 JSON으로 응답(키값은 반드시 fulfillmentText로)
    elif '브라우저' == program:
        import webbrowser
        webbrowser.open_new('https://www.google.com')
        reply = {'fulfillmentText': '브라우저 실행합니다.'}
    else:
        reply = {'fulfillmentText': 'can not excute {}'.format(program)}

    return jsonify(reply)

@app.route('/crawl',methods=['POST'])
def crawl():
    print('파이썬으로 왔따')



    print('필요한 영양소:',request.form['data'])
    driverPath = '{}\chromedriver.exe'.format(os.path.dirname(os.path.realpath(__file__)))

    options = webdriver.ChromeOptions()
    options.add_argument('headless')
    options.add_argument('window-size=1920x1080')
    options.add_argument('disable-gpu')
    options.add_argument(
        'User-Agent=Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36')

    driver = webdriver.Chrome(driverPath)
    driver_ = webdriver.Chrome(driverPath)#,options=options
    driver_3 = webdriver.Chrome(driverPath)
    driver.get('https://www.foodsafetykorea.go.kr/fcdb/detail/search/list.do')  # 자동으로 크롬 브라우저에 구글 페이지가 보인다

    try:
        # 음식 라벨 클릭
        element = driver.find_element_by_xpath('//*[@id="listForm"]/div[2]/div[1]/label[5]')
        element.click()
        # 영양소 단백질 클릭
        elements = driver.find_elements_by_xpath('//*[@id="nutri1List"]/li/a')
        for ele in elements:
            # print(ele.text)
            if ele.text == request.form['data']:
                ele.click()
        # 클릭
        driver.find_element_by_xpath('//*[@id="listForm"]/div[3]/ul[1]/li[1]/a').send_keys(Keys.ENTER)
        ele = WebDriverWait(driver, 5).until(EC.frame_to_be_available_and_switch_to_it((By.NAME, 'jqxIframe1')))
        a = WebDriverWait(driver, 5).until(EC.presence_of_element_located((By.XPATH, '//*[@id="depth1List"]/li[2]/a')))
        a.click()
        # 적용
        b = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="listForm"]/div[4]/a[1]')))
        b.click()
        c = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '//*[@id="listForm"]/div[4]/div/button')))
        c.click()
        d = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.XPATH, '// *[ @ id = "listForm"] / table / thead / tr / th[8] / span / a[2]')))
        d.click()




        elements_ = driver.find_elements_by_xpath('//*[@id="listForm"]/table/tbody/tr/td[2]/a')

        print(elements_[0].text.split(' ')[1])

        #크롤링한 메뉴로 이미지 가져오기
        get_img = 'https://www.google.com/search?q={}&tbm=isch'.format(quote_plus(elements_[0].text.split(' ')[1]))

        driver_3.get(get_img)

        image1 = WebDriverWait(driver_3, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#islrg > div.islrc > div:nth-child(1) > a.wXeWr.islib.nfEiy.mM5pbd > div.bRMDJf.islir > img')))
        image2 = WebDriverWait(driver_3, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#islrg > div.islrc > div:nth-child(2) > a.wXeWr.islib.nfEiy.mM5pbd > div.bRMDJf.islir > img')))
        image3 = WebDriverWait(driver_3, 5).until(EC.presence_of_element_located((By.CSS_SELECTOR, '#islrg > div.islrc > div:nth-child(3) > a.wXeWr.islib.nfEiy.mM5pbd > div.bRMDJf.islir > img')))

        image1_src=image1.get_attribute('src')
        image2_src = image2.get_attribute('src')
        image3_src = image3.get_attribute('src')





        #크롤링한 메뉴로 유트브 가져오기
        driver_.get('https://www.youtube.com/results?search_query={}'.format(elements_[0].text.split(' ')[1]))
        #driver_.find_element_by_xpath('//*[@id="search"]').send_keys(elements_[0].text.split(' ')[1])
        #driver_.find_element_by_xpath('//*[@id="search"]').send_keys(Keys.ENTER)
        videos = driver_.find_elements_by_xpath('//*[@id="thumbnail"]')

        href = videos[1].get_attribute('href')

        print(href)
        reply = {'food': '{}'.format(elements_[0].text.split(' ')[1]),'href':'{}'.format(href),'img1':'{}'.format(image1_src),'img2':'{}'.format(image2_src),'img3':'{}'.format(image3_src)}


    except TimeoutException as e:
        print('해당 페이지에 태그 요소가 존재하지 않거나,해당 페이지가 3초동안 열리지 않았어요:', e, sep='')
    finally:
        print("끝")


    return jsonify(reply)



if __name__ == '__main__':
    app.run(host='0.0.0.0',port=8383)